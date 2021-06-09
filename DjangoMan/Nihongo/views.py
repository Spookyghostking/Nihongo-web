from django.shortcuts import render
from django.http import HttpResponse, HttpResponseRedirect
from django.urls import reverse
import random
import openpyxl


class Answer:
    def __init__(self, list, num, correct=False):
        self.ans = list[num]
        self.ID = num
        self.correct = correct
        if correct:
            self.template = "<div class=correct>%s</div>"
        else:
            self.template = "<div class=incorrect>%s</div>"

    def __str__(self):
        return self.template % self.ans

class Dictionary:
    def __init__(self):
        # turn the excel entries into a python list
        sheetname = "Sheet1"

        wb = openpyxl.load_workbook("Nihongo .xlsx")
        datasheet = wb.get_sheet_by_name(sheetname)

        self.lengt = 0
        self.jap = ["This field is to remain empty for convenience purposes."]
        self.eng = ["This field is to remain empty for convenience purposes."]
        condition = True
        while condition:
            self.lengt += 1
            self.jap.append(datasheet.cell(column=2, row=self.lengt).value)
            self.eng.append(datasheet.cell(column=3, row=self.lengt).value)
            if datasheet.cell(column=2, row=self.lengt).value == None:
                condition = False
        self.jap.__delitem__(self.lengt)
        self.eng.__delitem__(self.lengt)
        self.lengt -= 1

        # for word in [datasheet.cell(column=2, row=x).value for x in range(1, self.lengt+1)]:
        #     self.jap.append(word)
        # for word in [datasheet.cell(column=3, row=x).value for x in range(1, self.lengt+1)]:
        #     self.eng.append(word)


def shuffle(list):
    for i in range(len(list)-1, -1, -1):
        j = random.randint(0, i)
        list[i], list[j] = list[j], list[i]
    return list

def generate_qlist(qr, inclusive=True):
    if inclusive: corr = 1
    else: corr = 0

    r = shuffle([i for i in range(qr[0], qr[1] + corr)])
    print([i for i in range(qr[0], qr[1] + corr)])
    return r

def randans(qnumber, sess, ld=Dictionary()):
    r = 0
    # make sure that there are no two same answers in the list
    while (r == 0) or (r == sess["qlist"][qnumber]) or (r in sess["alist"]):
        r = random.randint(1, ld.lengt-1)
    return r


# Create your views here.
def index(request):
    # if the site has not been visted before, start in the dead state
    if not "alive" in request.session:
        request.session["alive"] = False

    # If the request method is get, and a game is not in progress, render the page
    if request.method == "GET":
        if not request.session["alive"]:
            return render(request, "index.html")
        else: # if a game is in progress, redirect to the game site
            return HttpResponseRedirect(reverse("nihongo:flash"))
    elif request.method == "POST": # if the method is post, start new game
        try:
            qrange = (int(request.POST["begin"]), int(request.POST["end"]))
        except ValueError:
            return render(request, "index.html", {
                "message":"Please enter a valid range."
            })
        request.session["qlist"] = generate_qlist(qrange)

        request.session["alive"] = True
        return HttpResponseRedirect(reverse("nihongo:flash"))

def flash(request):
    # variable decleration
    ld = Dictionary()
    sess = request.session
    if not request.method == "POST":
        # begin new question
        # check if new question chain should begin
        if (not "question" in sess) or (sess["question"] == 0):
            sess["count"] = sess["remaining"] = len(sess["qlist"])
            sess["incorrect"] = 0
            sess["question"] = 1
        # extract question from session info
        try:
            qnumber = sess["question"] -1
            ambamba = sess["qlist"][qnumber]
        except IndexError: # if the end of the dictionary is reached, end question chain
            sess["alive"] = False
            sess["question"] = 0
            percent = 100 * sess["count"]/(sess["count"]+sess["incorrect"])
            return render(request, "index.html", {"message":f"{percent:.2f}%"})
        try:
            question = (ld.eng[ambamba], ld.jap[ambamba])
        except IndexError:
            sess["alive"] = False
            sess["question"] = 0
            # return HttpResponseRedirect(reverse("nihongo:index"))
            return render(request, "index.html", {"message":"Index out of range."})

        # generate a list of random answers
        generate_answers = True
        for ans in sess["alist"]:
            if ("class=correct" in ans) and (ld.eng[sess["qlist"][qnumber]] in ans):
                generate_answers = False
        if generate_answers:
            sess["alist"] = []
            for i in range(7):
                sess["alist"].append(randans(qnumber, sess, ld))
            # convert the list of answers into Answer objects
            sess["alist"] = [Answer(ld.eng, i) for i in sess["alist"]]
            # add the correct answer to the list, and then randomize
            acorrect = sess["qlist"][qnumber]
            acorrect = Answer(ld.eng, acorrect, correct=True)
            sess["alist"].append(acorrect)
            sess["alist"] = shuffle(sess["alist"])
            # convert answers to strings -_- (I thought this would be done automaticaly)
            sess["alist"] = [str(i) for i in sess["alist"]]
        return render(request, "flash.html", {
            "num":qnumber +1,
            "cou":sess["count"],
            "rem":sess["remaining"],
            "inc":sess["incorrect"],
            "eng":question[0],
            "jap":question[1],
            "ans":sess["alist"],
        })
    else:
        # end current question
        sess = request.session
        if "isCorrect" in request.POST:
            sess["question"] += 1
            if request.POST["isCorrect"] == "True":
                sess["remaining"] -= 1
                pass
            else:
                sess["qlist"].append(sess["qlist"][int(request.POST["qnum"])-1])
                sess["incorrect"] += 1

        return HttpResponseRedirect(reverse("nihongo:flash"))



def stop(request):
    # if request.method == "POST":
    request.session["alive"] = False
    request.session["question"] = 0
    return HttpResponseRedirect(reverse("nihongo:index"))
